from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import os
import pandas as pd
from django.core.files.storage import FileSystemStorage
import csv
from directories.models import *
from acts.models import *
from applications.models import *
from .models import loadsqls
from django.db import transaction
import subprocess
from django.db import connection
import openpyxl
from datetime import datetime
# from wand.image import Image as WandImage
# import pytesseract
# from PIL import Image


# pytesseract.pytesseract.tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'



@login_required
def Loadsqls(request):
    return render(request, 'loadsql/loadsql.html')


def get_table_columns(table_name):
    print('Получение столбцов из Excell...')
    with connection.cursor() as cursor:
        cursor.execute(f'PRAGMA table_info({table_name});')
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
    return column_names


def reorder_csv_columns(input_file, output_file, column_order):
    print('Изменение порядка столбцов CSV...')
    with open(input_file, 'r', encoding='utf-8') as infile:
        reader = csv.DictReader(infile)

        with open(output_file, 'w', encoding='utf-8', newline='') as outfile:
            writer = csv.DictWriter(outfile, fieldnames=column_order)
            writer.writeheader()

            for row in reader:
                # Преобразование даты
                old_date_str = row["Дата принятия к учету1"]
                old_date = datetime.strptime(old_date_str, "%d.%m.%Y %H:%M:%S")
                new_date_str = old_date.strftime("%Y-%m-%d %H:%M:%S")
                
                # Замена значения в столбце
                row["Дата принятия к учету1"] = new_date_str

                # Запись строки в выходной файл
                writer.writerow(row)


def import_csv_to_sqlite(csv_file_path, db_name, table_name):
    # Команда для выполнения импорта CSV в SQLite
    command = f'sqlite3 "{db_name}" ".mode csv" ".import {csv_file_path} {table_name}"'

    # Запуск команды в терминале
    subprocess.run(command, shell=True)


def upload_file(request):
    if request.method == 'POST':

        # Проверка на наличие файла в запросе
        if 'file' not in request.FILES:
            return HttpResponse("No file found in request", status=400)

        file = request.FILES['file']
        selected_option = request.POST.get('optionGroup', None)
        fs = FileSystemStorage(location='media/uploads')
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        file_extension = os.path.splitext(file.name)[1].lower()

        if file_extension != '.csv':
            if selected_option == 'IT_OS':
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active  # Выбираем активный лист

                column_number = None
                for col in sheet.iter_cols(min_row=1, max_row=1):
                    if col[0].value == "Основное средство.Инв. номер ДИТ":
                        # Получаем номер столбца
                        column_number = col[0].col_idx
                        break

                if column_number:
                    # Вставляем новый столбец сразу после найденного
                    sheet.insert_cols(column_number + 1)

                    # Дать имя новому столбцу
                    sheet.cell(row=1, column=column_number +
                               1).value = "Группа ОС"

                    # Проходимся по строкам
                    for row_num in range(2, sheet.max_row + 1):
                        cell_value = sheet.cell(
                            row=row_num, column=column_number).value
                        if cell_value is not None:
                            first_five_chars = cell_value[:5]

                            new_value = None
                            if first_five_chars == "ITEKS":
                                new_value = "Мини ПК ITEKS"
                            elif first_five_chars == "ITMNT":
                                new_value = "Мониторы ITMNT"
                            elif first_five_chars == "ITMNB":
                                new_value = "Моноблок ITMNB"
                            elif first_five_chars == "ITNTB":
                                new_value = "Ноутбук ITNTB"
                            elif first_five_chars == "ITPAD":
                                new_value = "Планшет ITPAD"
                            elif first_five_chars == "ITSRV":
                                new_value = "Сервер ITSRV"
                            elif first_five_chars == "ITSHD":
                                new_value = "Система хранения данных ITSHD"
                            elif first_five_chars == "ITWKS":
                                new_value = "Системный блок, Тонкий клиент ITWKS"
                            elif first_five_chars == "ITVDN":
                                new_value = "Видеорегистраторы ITVDN"
                            elif first_five_chars == "ITHDD":
                                new_value = "Жесткие диски ITHDD"
                            elif first_five_chars == "ITUPS":
                                new_value = "ИБП (источники бесперебойного питания) Стабилизатор ITUPS"
                            elif first_five_chars == "ITKSS":
                                new_value = "Кассовое оборудование ITKSS"
                            elif first_five_chars == "ITSAN":
                                new_value = "Оптические коммутаторы ITSAN"
                            elif first_five_chars == "ITVDC":
                                new_value = "Охранное видеонаблюдение (Видеокамеры ITVDC)"
                            elif first_five_chars == "ITPRN":
                                new_value = "Принтеры, МФУ, копировальные аппараты ITPRN"
                            elif first_five_chars == "ITPRK":
                                new_value = "Проектор ITPRK"
                            elif first_five_chars == "ITETH" or first_five_chars == "ITCSC":
                                new_value = "Сетевое оборудование ITETH"
                            elif first_five_chars == "ITSCN":
                                new_value = "Сканеры штрихкода ITSCN"
                            elif first_five_chars == "ITCNT" or first_five_chars == "ITKND":
                                new_value = "Счетчики посетителей ITCNT"
                            elif first_five_chars == "ITTLF":
                                new_value = "Телефон, факс ITTLF"
                            elif first_five_chars == "ITTCD":
                                new_value = "Терминал для сбора данных ITTCD,  подставки под ТСД, зарядные устройства для ТСД"
                            elif first_five_chars == "ITBOX":
                                new_value = "Шкаф коммутационный, серверный  ITBOX"

                        if new_value:
                            sheet.cell(
                                row=row_num, column=column_number + 1).value = new_value

                wb.save(file_path)
            try:
                # Чтение данных из Excel-файла
                data = pd.read_excel(file_path, engine='openpyxl')

                # Создание пути для нового CSV-файла
                csv_file_path = os.path.splitext(file_path)[0] + '.csv'

                # Сохранение данных в CSV-файл
                data.to_csv(csv_file_path, index=False, encoding='utf-8')

                # Удаление исходного файла, если необходимо
                os.remove(file_path)

                file_path = csv_file_path

            except Exception as e:
                return HttpResponse(f"An error occurred: {e}", status=400)
        if selected_option == 'Users':

            with transaction.atomic():
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as csvfile:
                    csvreader = csv.DictReader(csvfile)

                    for row in csvreader:
                        Users.objects.update_or_create(
                            name=row['ФИО'],
                            department=row['Департамент'],
                            position=row['Должность'],
                            organization=row['Организация'],
                            subdivision=row['Подразделение']

                        )
            loadsqls.objects.create(option=selected_option)
            os.remove(file_path)
            return redirect('/')

        elif selected_option == 'SkladyOffice':

            with transaction.atomic():
                with open(file_path, 'r', encoding='utf-8') as csvfile:
                    csvreader = csv.DictReader(csvfile)

                    for row in csvreader:
                        print(f"как бы: {row}")
                        SkladyOffice.objects.update_or_create(
                            sklad_name=row['Имя'],
                            sklad_type=row['Тип'],
                            sklad_city=row['Город'],
                            sklad_adress=row['Адрес'],
                        )
            loadsqls.objects.create(option=selected_option)
            os.remove(file_path)
            return redirect('/')

        elif selected_option == 'Tmc':

            with transaction.atomic():
                with open(file_path, 'r', encoding='utf-8') as csvfile:
                    csvreader = csv.DictReader(csvfile)

                    for row in csvreader:
                        Tmc.objects.update_or_create(
                            tmc_name=row['Номенклатура'],
                            tmc_article=row['Артикул'],
                            web_code=row['Web code'],
                            tmc_price=row['Себестоимость'],
                        )
            loadsqls.objects.create(option=selected_option)
            os.remove(file_path)
            return redirect('/directories/tmc')

        elif selected_option == 'IT_OS':

            column_name = get_table_columns(selected_option)
            mapping = {
                'inv_dit': 'Основное средство.Инв. номер ДИТ',
                'name_os': 'Основное средство.Наименование',
                'inpute_date': 'Дата принятия к учету1',
                'original_price': 'Стоимость для вычисления амортизации',
                'serial_number': 'Основное средство.Номер паспорта (регистрационный)',
                'os_group': 'Группа ОС',
            }
            column_order = []
            for i in range(len(column_name)):
                column_order.append(mapping.get(column_name[i], None))

            output_file_path = 'file.csv'
            reorder_csv_columns(file_path, output_file_path, column_order)

            db_name = 'db.sqlite3'
            import_csv_to_sqlite(output_file_path, db_name, selected_option)

            loadsqls.objects.create(option=selected_option)
            os.remove(file_path)
            os.remove(output_file_path)

            return redirect('/directories/os')

        return redirect('/loadsql')
    


def add_act_skans(request):
    pass
#     if request.method == 'POST':
#         uploaded_files = request.FILES.getlist('files')
        
#         for file in uploaded_files:
            
#             with open('media/uploads/acts/' + file.name, 'wb+') as destination:
#                 for chunk in file.chunks():
#                     destination.write(chunk)
        
#             file_name = os.path.splitext(file.name)
       
#             if file_name[1] == '.pdf' or file_name[1] == '.PDF':
       
#                 # Открываем PDF-файл и преобразуем его в JPEG
#                 pdf_path = 'media/uploads/acts/' + file.name
       
#                 with WandImage(filename=pdf_path, resolution=300) as pdf:
       
#                     for i, page in enumerate(pdf.sequence):
       
#                         with WandImage(page, resolution=300) as image:
       
#                             # Задаем имя и путь для сохранения JPEG-файла
#                             jpg_name = os.path.join('media/uploads/acts/', os.path.splitext(file.name)[0] + f'.page{i+1}.jpg')
       
#                             # Сохраняем JPEG-файл с высоким качеством
#                             image.compression_quality = 100  # максимальное качество
#                             image.save(filename=jpg_name)


            
#                 os.remove(pdf_path)

#         scans_names = []
#         for path in os.listdir('media/uploads/acts/'):
#             if 'Акт' in path:
#                 continue
#             elif os.path.isfile(os.path.join('media/uploads/acts/', path)):
#                 scans_names.append(path)

#         print(scans_names)

#         for i in scans_names:
#             file = os.path.join('media/uploads/acts/', i)
#             img = Image.open(file)

#             # Преобразуем изображение в текст
#             text = pytesseract.image_to_string(img, lang='rus + eng')
#             lines = text.split('\n')
#             print(lines)

#             if 'Акт' in lines[0]:
#                 a = "№"
#                 start_index = text.find(a + " ")
#                 result = text[start_index + len(a + " "):start_index + len(a + " ") + 3]
#                 result2 = result.replace('о', '0')
#                 result3 = 'Акт ТС №' + result2

#                 new_name = os.path.join('media/uploads/acts/', result3) + '.jpg'
#                 if os.path.exists(new_name):
#                     index = 1
#                     while True:
#                         new_name_parts = os.path.splitext(new_name)
#                         new_name = f"{new_name_parts[0]}.{index}{new_name_parts[1]}"
#                         if not os.path.exists(new_name):
#                             break
#                         index += 1

#                 # переименовываем файл
#                 os.rename(file, new_name)
            


#         return HttpResponse("Файлы успешно загружены.")
